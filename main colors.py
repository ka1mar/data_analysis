import openpyxl as op
from sklearn.cluster import KMeans
import cv2
#Кластеризация по цветовому тону (hue в hsv)

def get_dominant_color(image, k=3):
    image = image.reshape((image.shape[0] * image.shape[1], 3))
    clt = KMeans(n_clusters = k)
    clt.fit_predict(image)
    dominant_color = clt.cluster_centers_
    return list(dominant_color)

if __name__ == '__main__':
    wb = op.load_workbook(r'C:\Users\1626158\OneDrive\Рабочий стол\КУРСАЧ\cover_data.xlsx')
    ws = wb["1"]

    for id, row in enumerate(ws.iter_rows(min_row=2)):
        path = "images_copy/" + row[1].value
        print(path)
        cov = cv2.imread(path)
        hsv_image = cv2.cvtColor(cov, cv2.COLOR_RGB2HSV)
        dom_color = get_dominant_color(hsv_image)

        c = ws.cell(row=id + 2, column=14)
        c.value = str(int(dom_color[0][0]))
        c = ws.cell(row=id + 2, column=15)
        c.value = str(int(dom_color[1][0]))
        c = ws.cell(row=id + 2, column=16)
        c.value = str(int(dom_color[2][0]))

    wb.save(r'C:\Users\1626158\OneDrive\Рабочий стол\КУРСАЧ\cover_data.xlsx')
    print('DONE!')
