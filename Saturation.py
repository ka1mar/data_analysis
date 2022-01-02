import numpy as np
import openpyxl as op
import cv2
#перевод картинок в HSV-формат, анализ насыценности (первый и последний децили, медиана)

if __name__ == '__main__':
#    "D:\Sandbox\Github\opencv-project\Resources\Photos\cats.jpg"
    wb = op.load_workbook(r'C:\Users\1626158\OneDrive\Рабочий стол\КУРСАЧ\cover_data.xlsx')
    ws = wb["1"]

    for id, row in enumerate(ws.iter_rows(min_row=2)):
        path = "images_copy/" + row[1].value
        print(path)
        cov = cv2.imread(path)
        hsv_cov = cv2.cvtColor(cov, cv2.COLOR_RGB2HSV)
        h, s, v = cv2.split(hsv_cov)

        data = np.percentile(s, np.arange(10, 100, 10))  # deciles

        a = ws.cell(row = id+2, column=7)
        a.value = data[0]
        b = ws.cell(row=id + 2, column=8)
        b.value = data[4]
        c = ws.cell(row=id + 2, column=9)
        c.value = data[8]

    wb.save(r'C:\Users\1626158\OneDrive\Рабочий стол\КУРСАЧ\cover_data.xlsx')
