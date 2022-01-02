import openpyxl as op
import cv2
from mtcnn import MTCNN
#Применение MTCNN для нахождения лиц людей


if __name__ == '__main__':
    wb = op.load_workbook(r'C:\Users\1626158\OneDrive\Рабочий стол\КУРСАЧ\cover_data.xlsx')
    ws = wb["1"]
    detector = MTCNN()

    for id, row in enumerate(ws.iter_rows(min_row=2)):
        path = "images_copy/" + row[1].value
        cov = cv2.imread(path)
        c = ws.cell(row=id + 2, column=13)
        p = len(detector.detect_faces(cov))
        c.value = p
        print(p)
    print("ok")
    wb.save(r'C:\Users\1626158\OneDrive\Рабочий стол\КУРСАЧ\cover_data.xlsx')