import numpy as np
import openpyxl as op
import cv2
#перевод картинок в LAB-формат, анализ яркости (первый и последний децили, медиана)

if __name__ == '__main__':
#    "D:\Sandbox\Github\opencv-project\Resources\Photos\cats.jpg"
    wb = op.load_workbook(r'C:\Users\1626158\OneDrive\Рабочий стол\КУРСАЧ\cover_data.xlsx')
    ws = wb["1"]

    for id, row in enumerate(ws.iter_rows(min_row=2)):
        path = "images_copy/" + row[1].value
        print(path)
        cov = cv2.imread(path)
        lab_cov = cv2.cvtColor(cov, cv2.COLOR_RGB2LAB)[..., 0]

        data = np.percentile(lab_cov, np.arange(10, 100, 10))  # deciles

        a = ws.cell(row = id+2, column=4)
        a.value = data[0]
        b = ws.cell(row=id + 2, column=5)
        b.value = data[4]
        c = ws.cell(row=id + 2, column=6)
        c.value = data[8]

    wb.save(r'C:\Users\1626158\OneDrive\Рабочий стол\КУРСАЧ\cover_data.xlsx')
